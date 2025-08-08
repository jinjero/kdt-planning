import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
import json
import os
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Excel 파일 생성 함수
def create_excel_file(df, filename_prefix):
    """평가 템플릿 데이터를 3개 시트가 포함된 Excel 파일로 생성"""
    if not EXCEL_AVAILABLE:
        return False, "openpyxl 라이브러리가 설치되지 않았습니다. pip install openpyxl 명령어로 설치해주세요."
    
    try:
        # 새 워크북 생성
        wb = Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 시트 1: 출제자 평가 템플릿
        ws1 = wb.create_sheet("출제자 평가 템플릿")
        
        # 헤더 추가
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 추가
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row, 1):
                ws1.cell(row=row_num, column=col_num, value=value)
        
        # 시트 2: 출제자 문제 템플릿
        ws2 = wb.create_sheet("출제자 문제 템플릿")
        criteria_headers = ['대분류', '중분류', '소분류', '상 (90-100점)', '중 (70-89점)', '하 (50-69점)', '미달 (0-49점)', '비고']
        
        # 헤더 추가
        for col_num, header in enumerate(criteria_headers, 1):
            cell = ws2.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            # cell.fill = PatternFill(start_color="E67C73", end_color="E67C73", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 추가
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            criteria_row = [
                row.get('대분류', ''),
                row.get('중분류', ''),
                row.get('소분류', ''),
                row.get('상', ''),
                row.get('중', ''),
                row.get('하', ''),
                row.get('배점 X', ''),
                ''
            ]
            for col_num, value in enumerate(criteria_row, 1):
                ws2.cell(row=row_num, column=col_num, value=value)
        
        # 시트 3: 점수 집계표
        ws3 = wb.create_sheet("점수 집계표")
        score_headers = ['수험생명', '대분류', '중분류', '소분류', '배점', '획득점수', '평가자', '평가일시', '비고']
        
        # 헤더 추가
        for col_num, header in enumerate(score_headers, 1):
            cell = ws3.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 추가
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            score_row = [
                '',  # 수험생명
                row.get('대분류', ''),
                row.get('중분류', ''),
                row.get('소분류', ''),
                row.get('배점', ''),
                '',  # 획득점수
                '',  # 평가자
                '',  # 평가일시
                ''   # 비고
            ]
            for col_num, value in enumerate(score_row, 1):
                ws3.cell(row=row_num, column=col_num, value=value)
        
        # 모든 시트의 열 너비 자동 조정
        for ws in [ws1, ws2, ws3]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에서 Excel 파일 생성
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return True, {
            'data': excel_buffer.getvalue(),
            'filename': f"{filename_prefix}.xlsx"
        }
        
    except Exception as e:
        return False, f"Excel 파일 생성 중 오류가 발생했습니다: {str(e)}"

# CSV 다운로드 함수들
def create_csv_files(df, filename_prefix):
    """평가 템플릿 데이터를 3개의 CSV 파일로 생성"""
    try:
        csv_files = {}
        
        # 파일 1: 출제자 평가 템플릿
        template_csv = df.to_csv(index=False, encoding='utf-8-sig')
        csv_files['template'] = {
            'data': template_csv,
            'filename': f"{filename_prefix}_출제자평가템플릿.csv"
        }
        
        # 파일 2: 평가 기준표
        criteria_data = []
        for _, row in df.iterrows():
            criteria_row = {
                '대분류': row.get('대분류', ''),
                '중분류': row.get('중분류', ''),
                '소분류': row.get('소분류', ''),
                '상 (90-100점)': row.get('상', ''),
                '중 (70-89점)': row.get('중', ''),
                '하 (50-69점)': row.get('하', ''),
                '미달 (0-49점)': row.get('배점 X', ''),
                '비고': ''
            }
            criteria_data.append(criteria_row)
        
        criteria_df = pd.DataFrame(criteria_data)
        criteria_csv = criteria_df.to_csv(index=False, encoding='utf-8-sig')
        csv_files['criteria'] = {
            'data': criteria_csv,
            'filename': f"{filename_prefix}_평가기준표.csv"
        }
        
        # 파일 3: 점수 집계표
        score_data = []
        for _, row in df.iterrows():
            score_row = {
                '수험생명': '',
                '대분류': row.get('대분류', ''),
                '중분류': row.get('중분류', ''),
                '소분류': row.get('소분류', ''),
                '배점': row.get('배점', ''),
                '획득점수': '',
                '평가자': '',
                '평가일시': '',
                '비고': ''
            }
            score_data.append(score_row)
        
        score_df = pd.DataFrame(score_data)
        score_csv = score_df.to_csv(index=False, encoding='utf-8-sig')
        csv_files['score'] = {
            'data': score_csv,
            'filename': f"{filename_prefix}_점수집계표.csv"
        }
        
        return True, csv_files
        
    except Exception as e:
        return False, f"CSV 파일 생성 중 오류가 발생했습니다: {str(e)}"

# 구글 스프레드시트 API 설정
def setup_google_sheets():
    """구글 스프레드시트 API 설정"""
    try:
        credentials_path = os.getenv('GOOGLE_CREDENTIALS_PATH', 'credentials.json')
        if not os.path.exists(credentials_path):
            return None, "구글 서비스 계정 키 파일을 찾을 수 없습니다. credentials.json 파일을 업로드해주세요."
        
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        
        credentials = Credentials.from_service_account_file(credentials_path, scopes=scope)
        # Suppress all deprecation warnings from gspread
        import warnings
        warnings.filterwarnings("ignore", category=DeprecationWarning, module="gspread")
        client = gspread.authorize(credentials)
        return client, None
        
    except Exception as e:
        return None, f"구글 스프레드시트 API 설정 중 오류가 발생했습니다: {str(e)}"

# 구글 스프레드시트 관련 함수들
def export_to_google_sheets_with_templates(df, spreadsheet_name, folder_id=None):
    """데이터프레임을 3개 시트가 포함된 구글 스프레드시트로 내보내기"""
    client, error = setup_google_sheets()
    if error:
        return False, error
    
    try:
        # 스프레드시트 생성
        spreadsheet = client.create(spreadsheet_name)
        
        # 특정 폴더로 이동 (폴더 ID가 제공된 경우)
        if folder_id:
            try:
                # gspread를 통한 폴더 이동
                import googleapiclient.discovery
                from google.oauth2.service_account import Credentials
                
                # Drive API 서비스 생성
                credentials = client.auth
                drive_service = googleapiclient.discovery.build('drive', 'v3', credentials=credentials)
                
                # 폴더 존재 여부 확인
                try:
                    drive_service.files().get(fileId=folder_id).execute()
                except:
                    print(f"폴더 ID {folder_id}를 찾을 수 없습니다. 기본 위치에 생성됩니다.")
                    folder_id = None
                
                if folder_id:
                    # 현재 부모 폴더 가져오기
                    file = drive_service.files().get(fileId=spreadsheet.id, fields='parents').execute()
                    previous_parents = ",".join(file.get('parents'))
                    
                    # 새 폴더로 이동
                    drive_service.files().update(
                        fileId=spreadsheet.id,
                        addParents=folder_id,
                        removeParents=previous_parents,
                        fields='id, parents'
                    ).execute()
                
            except Exception as e:
                # 폴더 이동 실패해도 스프레드시트 생성은 계속 진행
                print(f"폴더 이동 중 오류 (스프레드시트는 정상 생성됨): {e}")
        
        # 기본 시트 삭제 (Sheet1)
        try:
            default_sheet = spreadsheet.sheet1
            spreadsheet.del_worksheet(default_sheet)
        except:
            pass
        
        # 시트 1: 출제자 평가 템플릿
        sheet1 = spreadsheet.add_worksheet(title="출제자 평가 템플릿", rows=1000, cols=20)
        template_data = [df.columns.tolist()] + df.values.tolist()
        sheet1.update(range_name='A1', values=template_data)
        
        # 시트 1 헤더 서식 적용
        sheet1.format('A1:I1', {
            'backgroundColor': {'red': 0.2, 'green': 0.6, 'blue': 0.9},
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}}
        })
        
        # 시트 2: 평가 기준표
        sheet2 = spreadsheet.add_worksheet(title="평가 기준표", rows=1000, cols=20)
        criteria_headers = ['대분류', '중분류', '소분류', '상 (90-100점)', '중 (70-89점)', '하 (50-69점)', '미달 (0-49점)', '비고']
        criteria_data = [criteria_headers]
        
        # 기존 데이터에서 평가 기준 추출
        for _, row in df.iterrows():
            criteria_row = [
                row.get('대분류', ''),
                row.get('중분류', ''),
                row.get('소분류', ''),
                row.get('상', ''),
                row.get('중', ''),
                row.get('하', ''),
                row.get('배점 X', ''),
                ''
            ]
            criteria_data.append(criteria_row)
        
        sheet2.update(range_name='A1', values=criteria_data)
        sheet2.format('A1:H1', {
            'backgroundColor': {'red': 0.9, 'green': 0.6, 'blue': 0.2},
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}}
        })
        
        # 시트 3: 점수 집계표
        sheet3 = spreadsheet.add_worksheet(title="점수 집계표", rows=1000, cols=20)
        score_headers = ['수험생명', '대분류', '중분류', '소분류', '배점', '획득점수', '평가자', '평가일시', '비고']
        score_data = [score_headers]
        
        # 각 평가 항목별로 점수 입력 행 생성
        for _, row in df.iterrows():
            score_row = [
                '',  # 수험생명 (입력 필요)
                row.get('대분류', ''),
                row.get('중분류', ''),
                row.get('소분류', ''),
                row.get('배점', ''),
                '',  # 획득점수 (입력 필요)
                '',  # 평가자 (입력 필요)
                '',  # 평가일시 (입력 필요)
                ''   # 비고
            ]
            score_data.append(score_row)
        
        sheet3.update(range_name='A1', values=score_data)
        sheet3.format('A1:I1', {
            'backgroundColor': {'red': 0.6, 'green': 0.9, 'blue': 0.2},
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}}
        })
        
        # 스프레드시트를 공개로 설정
        spreadsheet.share('', perm_type='anyone', role='reader')
        
        # 스프레드시트 URL 반환
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet.id}"
        return True, spreadsheet_url
        
    except Exception as e:
        return False, f"스프레드시트 내보내기 중 오류가 발생했습니다: {str(e)}"

def export_to_google_sheets(df, spreadsheet_name, worksheet_name="출제자 평가 템플릿"):
    """기존 단일 시트 내보내기 함수 (호환성 유지)"""
    return export_to_google_sheets_with_templates(df, spreadsheet_name)

# 페이지 설정
st.set_page_config(
    page_title="스파르타 취업 역량 평가",
    page_icon="🧑🏻‍🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 메인 타이틀
st.title("🧑🏻‍🎓 스파르타 취업 역량 평가")
st.markdown("---")

# 세션 상태 초기화
if 'current_page' not in st.session_state:
    st.session_state.current_page = "평가표"

# 버튼 메뉴
if st.sidebar.button("평가표", use_container_width=True):
    st.session_state.current_page = "평가표"
if st.sidebar.button("출제자 평가 템플릿", use_container_width=True):
    st.session_state.current_page = "출제자 평가 템플릿"
if st.sidebar.button("출제자 문제 템플릿", use_container_width=True):
    st.session_state.current_page = "출제자 문제 템플릿"

page = st.session_state.current_page

if page == "평가표":
    st.header("평가표")

    st.info("""
        **📋 평가 템플릿 생성 방법**
        1. 개발/비개발 과목을 선택
        2. 원하는 평가 항목들을 선택
        3. "평가 템플릿에 추가" 버튼 클릭
        4. "출제자 평가 템플릿" 페이지로 이동해 내용을 편집하고 구글 스프레드시트로 내보내기
    """)
    
    # 개발/비개발 토글
    job_type = st.selectbox("과목 선택", ["개발", "비개발"])
    
    if job_type == "개발":
        st.subheader("개발 평가표")
        # 계층 구조 정의 (예시)
        dev_hierarchy = {
            '코드': {
                '요구사항 해석': ['프로젝트 과제 목적에 맞게 구현했는가', '제공된 입출력 데이터를 통과하는가'],
                '알고리즘/로직': ['기본 문법(반복문, 조건문 등) 적절히 활용했는가', '효율적인 알고리즘 및 자료구조 사용했는가'],
                '코드 품질 및 최적화': ['적절한 함수 분리 및 모듈화가 되어있는가', '시간/공간복잡도 개선 노력이 있었는가', '중복 코드를 최소화하여 간결하게 작성했는가'],
                '예외 처리': ['예상 가능한 예외 사항을 고려했는가', 'try-catch-finally 구문을 적절히 사용했는가']
            },
            '프레임워크': {
                '구조 설계 이해도': ['프레임워크 구조 특성을 이해하고 적용했는가', '디렉토리 구조를 일관성있게 설계했는가'],
                '기능 구현 방식 적절성': ['프레임워크 방식에 맞춰 기능을 구현했는가', '내장 기능과 라이브러리를 알맞게 활용했는가'],
                '역할 분리 및 재사용성': ['로직, 서비스 등을 목적에 따라 분리했는가', '컴포넌트화, 모듈화 등을 통해 재사용을 할 수 있는가'],
                '상태 및 흐름 관리': ['상태 관리나 요청-응답 흐름을 일관되게 처리했는가', '프레임워크에 맞는 상태/라우팅 방식을 사용했는가'],
                '설정 및 의존성 관리': ['환경설정 파일(.env, web.xml 등)을 적절히 구성했는가', '외부 라이브러리, 모듈 의존성을 관리했는가'],
            }
        }
        # 표로 변환
        rows = []
        for 대분류, 중분류_dict in dev_hierarchy.items():
            for 중분류, 소분류_list in 중분류_dict.items():
                for 소분류 in 소분류_list:
                    rows.append({'대분류': 대분류, '중분류': 중분류, '소분류': 소분류})
        df = pd.DataFrame(rows)

    else:
        st.subheader("비개발 평가표")
        # 계층 구조 정의 (예시)
        biz_hierarchy = {
            '기획': {
                '문제 정의': ['해결해야 할 문제와 핵심 이슈를 제대로 설정했는가'],
                '요구사항 분석': ['고객/시장/업무의 요구사항을 잘 분석했는가', '다양한 관계자의 니즈를 반영했는가', '충돌되는 요구사항을 조율했는가'],
                '목표 설정': ['달성 가능한 목표/지표를 설계했는가', '핵심 기능 또는 가치 요소를 제대로 도출했는가'],
                '전략 및 기획': ['목표를 달성하기 위한 구체적인 전략을 수립했는가', '전개 방식이 일관되고 설득력있는 전략/기획인가'],
            },
            '완성도': {
                '결과물 완성도': ['계획한 목표를 달성했는가', '목표 달성 과정이 설득력있게 정리되었는가'],
                '문제 해결력': ['정성적/정량적 데이터를 적절히 활용했는가', '데이터 해석이 설득력을 갖고 판단의 근거로 기능했는가'],
                '전문성': ['직무에 맞는 툴을 목적에 맞게 활용했는가', '직무 용어 및 개념을 올바르게 사용했는가'],
                '성과 분석': ['결과물에 대한 성과 분석을 수행했는가', '문제점과 긍정적인 성과 모두를 도출했는가'],
                '개선 제안': ['수행 과정을 바탕으로 개선 방향을 논리적으로 제시했는가']
            },
            '소프트스킬': {
                '협업 및 전달력': ['다른 직무 담당자와의 협업을 고려했는가', '기획 의도, 결과물을 명확하게 설명했는가'],
                '창의성': ['결과물을 도출하기 위한 과정이 창의적으로 진행됐는가', '다른 수험생들과 비교되는 지점이 있는가'],
            }
        }
        # 표로 변환
        rows = []
        for 대분류, 중분류_dict in biz_hierarchy.items():
            for 중분류, 소분류_list in 중분류_dict.items():
                for 소분류 in 소분류_list:
                    rows.append({'대분류': 대분류, '중분류': 중분류, '소분류': 소분류})
        df = pd.DataFrame(rows)

    
    # 평가 표 편집 가능하게 표시
    # 드롭다운 옵션 준비
    dev_major_options = list(dev_hierarchy.keys()) if job_type == "개발" else []
    dev_mid_options = sum([list(v.keys()) for v in dev_hierarchy.values()], []) if job_type == "개발" else []
    dev_sub_options = sum([sum([vv for vv in v.values()], []) for v in dev_hierarchy.values()], []) if job_type == "개발" else []
    biz_major_options = list(biz_hierarchy.keys()) if job_type == "비개발" else []
    biz_mid_options = sum([list(v.keys()) for v in biz_hierarchy.values()], []) if job_type == "비개발" else []
    biz_sub_options = sum([sum([vv for vv in v.values()], []) for v in biz_hierarchy.values()], []) if job_type == "비개발" else []

    # 각 표에 맞는 옵션 지정
    if job_type == "개발":
        col_config = {
            "대분류": st.column_config.SelectboxColumn("대분류", options=dev_major_options, required=True),
            "중분류": st.column_config.SelectboxColumn("중분류", options=dev_mid_options, required=True),
            "소분류": st.column_config.TextColumn("소분류", width="large"),
        }
    else:
        col_config = {
            "대분류": st.column_config.SelectboxColumn("대분류", options=biz_major_options, required=True),
            "중분류": st.column_config.SelectboxColumn("중분류", options=biz_mid_options, required=True),
            "소분류": st.column_config.TextColumn("소분류", width="large"),
        }

    st.dataframe(df, use_container_width=True)

    # 멀티셀렉트로 행 선택 (인덱스 기준)
    selected_idx = st.multiselect(
        "추가할 행(들)을 선택하세요",
        options=df.index,
        format_func=lambda x: f"{x+1}행: {df.loc[x, '대분류']} / {df.loc[x, '중분류']} / {df.loc[x, '소분류']}"
    )

    # 9개 열 구성 (앞 3개: 대분류, 중분류, 소분류)
    template_columns = [
        "대분류", "중분류", "소분류",
        "평가 내용", "배점", "상", "중", "하", "배점 X"
    ]
    if st.button("평가 템플릿에 추가"):
        if selected_idx:
            selected = df.loc[selected_idx]
            selected_template = pd.DataFrame({
                "대분류": selected["대분류"].values,
                "중분류": selected["중분류"].values,
                "소분류": selected["소분류"].values,
                "평가 내용": "",
                "배점": "",
                "상": "",
                "중": "",
                "하": "",
                "배점 X": ""
            })
            if "template_table" not in st.session_state:
                st.session_state["template_table"] = pd.DataFrame(columns=template_columns)
            st.session_state["template_table"] = pd.concat([
                st.session_state["template_table"], selected_template
            ], ignore_index=True).drop_duplicates()
            st.success("평가 템플릿에 추가되었습니다!")
        else:
            st.warning("추가할 행을 먼저 선택해 주세요.")

elif page == "출제자 평가 템플릿":
    st.header("출제자 평가 템플릿")
    
    # 트랙명 입력 필드
    track_name = st.text_input(
        "트랙명",
        value="",
        placeholder="예: PM, UXUI, 그래픽 디자이너, 실시간 커머스, 디지털 마케터 등",
    )
    
    # 구글 드라이브 관련 UI 임시 비활성화
    # with st.expander("🗂️ 구글 드라이브 폴더 URL"):
    #     folder_url = st.text_input(
    #         "구글 드라이브 폴더 URL",
    #         value="",
    #         placeholder="구글 드라이브에서 폴더 URL을 복사해서 붙여넣으세요."
    #     )
    #     
    #     # URL에서 폴더 ID 추출
    #     folder_id = None
    #     if folder_url.strip():
    #         import re
    #         # 구글 드라이브 폴더 URL 패턴 매칭
    #         pattern = r'folders/([a-zA-Z0-9-_]+)'
    #         match = re.search(pattern, folder_url)
    #         if match:
    #             folder_id = match.group(1)
    #             st.success(f"✅ 폴더 ID 추출 성공: {folder_id}")
    #         else:
    #             st.error("❌ 올바른 구글 드라이브 폴더 URL을 입력해주세요.")
    
    if "template_table" in st.session_state:
        # 텍스트 입력 가능한 컬럼 설정
        col_config = {
            "평가 내용": st.column_config.TextColumn("평가 내용", width="medium"),
            "배점": st.column_config.NumberColumn("배점", width="small", format="%d"),
            "상": st.column_config.TextColumn("상", width="medium"),
            "중": st.column_config.TextColumn("중", width="medium"),
            "하": st.column_config.TextColumn("하", width="medium"),
            "배점 X": st.column_config.TextColumn("배점 X", width="medium"),
        }
        
        # 데이터 편집기
        edited_df = st.data_editor(
            st.session_state["template_table"],
            column_config=col_config,
            use_container_width=True,
            num_rows="dynamic"
        )
        
        # 편집된 데이터를 세션 상태에 저장
        st.session_state["template_table"] = edited_df
        
        # 파일 다운로드 섹션
        st.markdown("---")
        st.subheader("📥 파일 다운로드")
        
        # 다운로드 옵션 탭
        tab1, tab2, tab3 = st.tabs(["🌐 구글 스프레드시트", "📊 Excel 파일 (3개 시트)", "📄 CSV 파일 (개별)"])
        
        # 파일명 설정
        if track_name.strip():
            default_name = f"{track_name.strip()}_{datetime.now().strftime('%y%m%d')}"
        else:
            default_name = f"트랙명_{datetime.now().strftime('%y%m%d')}"
        
        filename_prefix = st.text_input(
            "파일명", 
            value=default_name,
            help="다운로드될 파일의 이름입니다. 트랙명을 입력하면 자동으로 업데이트됩니다."
        )
        
        with tab1:
            st.markdown("🌐 **구글 스프레드시트로 내보내기**")
            st.markdown("하나의 스프레드시트에 3개 시트가 포함됩니다:")
            st.markdown("- 시트 1: 출제자 평가 템플릿")
            st.markdown("- 시트 2: 평가 기준표")
            st.markdown("- 시트 3: 점수 집계표")
            
            # 구글 드라이브 폴더 URL 입력
            folder_url = st.text_input(
                "구글 드라이브 폴더 URL (선택사항)",
                placeholder="https://drive.google.com/drive/folders/1abcd...",
                help="스프레드시트를 저장할 구글 드라이브 폴더 URL을 입력하세요. 비워두면 내 드라이브 루트에 생성됩니다."
            )
            
            # 폴더 ID 추출
            folder_id = None
            if folder_url.strip():
                try:
                    if "/folders/" in folder_url:
                        folder_id = folder_url.split("/folders/")[1].split("?")[0]
                        st.success(f"✅ 폴더 ID: {folder_id}")
                    else:
                        st.warning("⚠️ 올바른 구글 드라이브 폴더 URL을 입력해주세요.")
                except:
                    st.error("❌ URL에서 폴더 ID를 추출할 수 없습니다.")
            
            # 구글 스프레드시트 내보내기 버튼
            google_button = st.button(
                "🌐 구글 스프레드시트로 내보내기",
                type="primary",
                use_container_width=True,
                key="google_export"
            )
            
            if google_button:
                if len(edited_df) == 0:
                    st.warning("내보낼 데이터가 없습니다.")
                else:
                    with st.spinner("구글 스프레드시트를 생성하는 중..."):
                        success, result = export_to_google_sheets_with_templates(
                            edited_df, 
                            filename_prefix, 
                            folder_id
                        )
                        
                        if success:
                            st.success("✅ 구글 스프레드시트를 성공적으로 생성했습니다!")
                            st.markdown(f"[🔗 스프레드시트 열기]({result})")
                            
                            st.info("""
                            📁 **생성된 시트 정보:**
                            - **시트 1: 출제자 평가 템플릿** - 기본 평가 항목 및 내용
                            - **시트 2: 평가 기준표** - 상/중/하 평가 기준 상세 정보
                            - **시트 3: 점수 집계표** - 실제 평가 시 사용할 점수 입력표
                            """)
                        else:
                            st.error(f"❌ 구글 스프레드시트 생성 실패: {result}")
                            
                            # 오류 해결 방법 안내
                            with st.expander("🔧 문제 해결 방법"):
                                st.markdown("""
                                **주요 오류 해결 방법:**
                                1. `credentials.json` 파일이 프로젝트 루트에 있는지 확인
                                2. 구글 서비스 계정에 Google Sheets API와 Google Drive API가 활성화되어 있는지 확인
                                3. 서비스 계정에 충분한 권한이 있는지 확인
                                4. 인터넷 연결 상태 확인
                                """)
        
        with tab2:
            st.markdown("하나의 Excel 파일에 3개 시트가 포함됩니다:")
            st.markdown("- 시트 1: 출제자 평가 템플릿")
            st.markdown("- 시트 2: 평가 기준표")
            st.markdown("- 시트 3: 점수 집계표")
            
            if EXCEL_AVAILABLE:
                excel_button = st.button(
                    "📊 Excel 파일 다운로드", 
                    type="primary",
                    use_container_width=True,
                    key="excel_download"
                )
                
                if excel_button:
                    if len(edited_df) == 0:
                        st.warning("다운로드할 데이터가 없습니다.")
                    else:
                        with st.spinner("Excel 파일을 생성하는 중..."):
                            success, result = create_excel_file(edited_df, filename_prefix)
                            
                            if success:
                                st.success("✅ Excel 파일을 성공적으로 생성했습니다!")
                                
                                st.download_button(
                                    label="📊 Excel 파일 다운로드",
                                    data=result['data'],
                                    file_name=result['filename'],
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    use_container_width=True
                                )
                                
                                st.info("""
                                📁 **Excel 파일 정보:**
                                - **시트 1: 출제자 평가 템플릿** - 기본 평가 항목 및 내용
                                - **시트 2: 평가 기준표** - 상/중/하 평가 기준 상세 정보
                                - **시트 3: 점수 집계표** - 실제 평가 시 사용할 점수 입력표
                                """)
                            else:
                                st.error(f"❌ Excel 파일 생성 실패: {result}")
            else:
                st.error("❌ Excel 기능을 사용하려면 openpyxl 라이브러리가 필요합니다.")
                st.code("pip install openpyxl")
        
        with tab3:
            st.markdown("**3개의 개별 CSV 파일로 다운로드:**")
            
            csv_button = st.button(
                "📄 CSV 파일들 다운로드", 
                type="secondary",
                use_container_width=True,
                key="csv_download"
            )
        
            # CSV 다운로드 버튼 클릭 시
            if csv_button:
                if len(edited_df) == 0:
                    st.warning("다운로드할 데이터가 없습니다.")
                else:
                    with st.spinner("3개의 CSV 파일을 생성하는 중..."):
                        success, csv_files = create_csv_files(edited_df, filename_prefix)
                        
                        if success:
                            st.success("✅ 3개의 CSV 파일을 성공적으로 생성했습니다!")
                            
                            # 각 CSV 파일에 대한 다운로드 버튼 생성
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                st.download_button(
                                    label="📋 출제자 평가 템플릿",
                                    data=csv_files['template']['data'],
                                    file_name=csv_files['template']['filename'],
                                    mime='text/csv',
                                    use_container_width=True
                                )
                            
                            with col2:
                                st.download_button(
                                    label="📊 평가 기준표",
                                    data=csv_files['criteria']['data'],
                                    file_name=csv_files['criteria']['filename'],
                                    mime='text/csv',
                                    use_container_width=True
                                )
                            
                            with col3:
                                st.download_button(
                                    label="📈 점수 집계표",
                                    data=csv_files['score']['data'],
                                    file_name=csv_files['score']['filename'],
                                    mime='text/csv',
                                    use_container_width=True
                                )
                            
                            # 파일 정보 표시
                            st.info("""
                            📁 **생성된 파일 정보:**
                            - **출제자 평가 템플릿** - 기본 평가 항목 및 내용
                            - **평가 기준표** - 상/중/하 평가 기준 상세 정보
                            - **점수 집계표** - 실제 평가 시 사용할 점수 입력표
                            """)
                        else:
                            st.error(f"❌ CSV 파일 생성 실패: {csv_files}")

    else:
        st.info("아직 추가된 항목이 없습니다. 먼저 '평가표' 페이지에서 항목을 추가해주세요.")
        
        # openpyxl 설치 안내
        if not EXCEL_AVAILABLE:
            st.warning("💡 Excel 파일 기능을 사용하려면 openpyxl 라이브러리 설치가 필요합니다.")
            st.code("pip install openpyxl")

# 푸터
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center'>
        <p>Copyright ⓒ TeamSparta All rights reserved.</p>
    </div>
    """,
    unsafe_allow_html=True
)
